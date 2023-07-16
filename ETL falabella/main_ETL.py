import pandas as pd
import os
import numpy as np
import time
from datetime import datetime

from openpyxl import Workbook

from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.worksheet.table import TableStyleInfo
from openpyxl.worksheet.table import Table
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter





now = datetime.now()


dir_actual = os.path.dirname(__file__)

# concatenar ruta del archivo de Excel




# PARA CORRERLO SE ENCESITAN 3 ARCHIVOS, "wip.csv" QUE ES EL INFORME DE RUTEO CONJUNTO, "preruteo.csv", QUE ES LA GRILLA MAPEADA DE SODIMAC, y "PREGRILLA.XLSX" CON LA HOJA DE LOS DATOS LLAMADA "1" 
# QUE ES LA GRILLA DE CAMILA

df_wip= pd.read_csv(os.path.join(dir_actual, "wip.csv"))
df_sodimac = pd.read_csv(os.path.join(dir_actual, "preruteo.csv"), encoding='ISO-8859-1')
#df_geosort =pd.read_csv(os.path.join(dir_actual, "geo.csv"), sep=";")
df_grilla = pd.read_excel(os.path.join(dir_actual, 'pregrilla.xlsx'),sheet_name="1",dtype={"Lpn": str})
df_flujo = pd.read_excel(os.path.join(dir_actual, "flujo.xlsx"), sheet_name="Page2", header=1) #
#df_muertos = pd.read_csv(os.path.join(dir_actual, "muertos.txt"),header=None, names=["LPN"])
df_falabella = pd.read_excel(os.path.join(dir_actual, 'falabella.xlsx'),sheet_name="1",dtype={'Lpn': str})



#CREACION Y EXTRACCION DE RESERVAS DE SODIMAC DE LA GRILAL DE FALABELLA
df_falabella = df_falabella.loc[df_falabella['Bu'] == 'Sodimac']
df_falabella = df_falabella[["Suborden","Producto","Bu","Lpn","Group_by","Direccion","Localidad","Patente"]]
df_falabella = df_falabella.rename(columns={'Lpn': 'LPN'})
df_falabella = df_falabella.rename(columns={'Group_by': 'RESERVA'})

df_falabella['RESERVA'] = df_falabella['RESERVA'].astype(str)
df_falabella['LPN'] = df_falabella['LPN'].astype(str)
df_falabella['LPN'] = df_falabella['LPN'].str.rstrip()
df_falabella['Suborden'] = df_falabella['Suborden'].astype(str).str.replace(".0","",regex=False)
df_falabella = df_falabella[["Suborden","Producto","Bu","LPN","RESERVA","Direccion","Localidad","Patente"]]





#extracion de data del flujo para procesamiento con las grillas
df_flujo = df_flujo[["TC_ORDER_ID","LPN","OC","DESTINO","ORIGEN","ULTIMA_FECHA_ACT_UB"]]
df_flujo = df_flujo.loc[df_flujo["TC_ORDER_ID"].astype(str).str.startswith("DOCSM").astype(bool)]
df_flujo["LPN"] = df_flujo["LPN"].astype(str)
df_flujo["ORIGEN"] = df_flujo["ORIGEN"].astype(int)
mask = df_flujo["ORIGEN"] == 9628
df_flujo = df_flujo[mask]





# contar direcciones unicas por patente grilla de sodimac-falabella
df_direcciones = df_grilla.groupby('Patente')['Direccion'].nunique().reset_index(name='DIRECCIONES UNICAS')

# concatenar localidades por patente grilla de sodimac-falabella
df_localidades = df_grilla.groupby('Patente')['Localidad'].unique().reset_index(name='LOCALIDADES')
df_localidades['LOCALIDADES'] = df_localidades['LOCALIDADES'].apply(lambda x: ', '.join(x))

# contar registros por patente grilla de sodimac-falabella
df_count = df_grilla.groupby('Patente').size().reset_index(name='LPN')

# unir dataframes
df_final = df_direcciones.merge(df_localidades, on='Patente').merge(df_count, on='Patente')

df_final = df_final.reindex(columns =["Patente","LPN","DIRECCIONES UNICAS","LOCALIDADES"])






df_wip['FECHA_PACTADA'] = pd.to_datetime(df_wip['FECHA_PACTADA'], format='%b %d, %Y')

df_wip['DO'] = df_wip['DO'].astype(str)
df_wip['SKU'] = df_wip['SKU'].astype(str)
df_wip['LPN'] = df_wip['LPN'].astype(str)
df_wip['SOC'] = df_wip['SOC'].astype(str).str.replace(".0","",regex=False)

df_wip['DO'] = df_wip['DO'].str.rstrip()
df_wip['SKU'] = df_wip['SKU'].str.rstrip()
df_wip['LPN'] = df_wip['LPN'].str.rstrip()
#df_wip['SOC'] = df_wip['SOC'].str.rstrip()



df_wip= df_wip.rename(columns={'DO': 'RESERVA'})

# Agregar una nueva columna con la concatenación de "DO" y "SKU"
df_wip.insert(loc=2, column='LLAVE', value=df_wip['RESERVA'].astype(str) + '-' + df_wip['SKU'].astype(str))

df_wip = df_wip.reindex(columns=["LPN","PATENTE_SHIP","RESERVA","SKU","PRODUCTO","CANTIDAD","LLAVE","SOC","SHIPMENT","ORIGEN","DESTINO","FECHA_PACTADA","METODO_DESPACHO",
                                 "VOLUMEN","TAMANO","REGION","COMUNA","DIRECCION","CLIENTE","FONO","DETALLE_STD"])

#print(df_wip)






df_sodimac['RESERVA'] = df_sodimac['RESERVA'].astype(str)
df_sodimac['SKU'] = df_sodimac['SKU'].astype(str)

df_sodimac['RESERVA'] = df_sodimac['RESERVA'].str.rstrip()
df_sodimac['SKU'] = df_sodimac['SKU'].str.rstrip()

df_sodimac = df_sodimac.rename(columns={'VOLUMEN': 'VOLUMEN_SODI'})
#crea columna llave
df_sodimac.insert(loc=1, column='LLAVE', value=df_sodimac['RESERVA'].astype(str) + '-' + df_sodimac['SKU'].astype(str))
#elimina las llaves duplicadas preservando el ultimo valor duplicado, que por lo general es "C/D"
df_sodimac = df_sodimac.drop_duplicates(subset="LLAVE",keep="last")


df_sodimac['PESO'] = df_sodimac['PESO'].astype(float)
df_sodimac['VOLUMEN_SODI'] = df_sodimac['VOLUMEN_SODI'].astype(float)







df_wip = df_wip.merge(df_sodimac[['LLAVE', 'JAULA','DESP','BLQ','ESTADO RESERVA','TIPO DE RESERVA','PESO','VOLUMEN_SODI']], on='LLAVE', how='left')
df_wip = df_wip.rename(columns={'DESP': 'BOD'})
df_wip = df_wip.rename(columns={'BLQ': 'BLOQUEADO'})
df_wip = df_wip.rename(columns={'ESTADO RESERVA': 'ESTADO SODIMAC'})

df_wip['LPN'] = df_wip['LPN'].astype(str)




df_grilla['Suborden'] = df_grilla['Suborden'].astype(str)
df_grilla['Suborden'] = df_grilla['Suborden'].str.rstrip()
df_grilla['Lpn'] = df_grilla['Lpn'].astype(str)
df_grilla['Lpn'] = df_grilla['Lpn'].str.rstrip()

df_grilla['Group_by'] = df_grilla['Group_by'].astype(str)
df_grilla['Group_by'] = df_grilla['Group_by'].str.rstrip()

df_grilla_resumen = df_grilla[["Posruta","Lpn","Suborden", "Group_by", "Producto","Direccion","Localidad","Patente","Nombrecliente","Rutcliente","Idruta"]]
df_grilla_resumen = df_grilla_resumen.rename(columns={'Suborden': 'SOC'})
df_grilla_resumen = df_grilla_resumen.rename(columns={'Group_by': 'RESERVA'})
df_grilla_resumen = df_grilla_resumen.rename(columns={'Lpn': 'LPN'})


df_grilla_resumen['LPN'] = df_grilla_resumen['LPN'].astype(str)







df_grilla_resumen = df_grilla_resumen.merge(df_wip[['LPN','SKU','CANTIDAD', 'JAULA','BOD','BLOQUEADO','FONO','ESTADO SODIMAC','TIPO DE RESERVA','DETALLE_STD']], on='LPN', how='left')





df_grilla_resumen["JAU/STD/ID"] = df_grilla_resumen["JAULA"].astype(str).str.replace(".0","",regex=False) + " // " + df_grilla_resumen["TIPO DE RESERVA"].astype(str) + " // " + df_grilla_resumen["Idruta"].astype(str) 

df_grilla_resumen["RESER/SOC/SKU"] = df_grilla_resumen["RESERVA"].astype(str) + " // " + df_grilla_resumen["SOC"].astype(str) + " // " + df_grilla_resumen["SKU"].astype(str)
df_grilla_resumen["COMUNA/PAT/EST"] =df_grilla_resumen["Localidad"].astype(str) + " // " + df_grilla_resumen["Patente"].astype(str) + " // " + df_grilla_resumen["DETALLE_STD"].astype(str)
df_grilla_resumen["CLIENTE/TELEF"] = df_grilla_resumen["Nombrecliente"].astype(str) + " // " + df_grilla_resumen["FONO"].astype(str)

empty_series = pd.Series(index=df_grilla_resumen.index, dtype=object)
#

# agrega la serie vacía como una nueva columna al DataFrame

df_grilla_resumen["V/M"] = empty_series
df_grilla_resumen["BULTOS"] = empty_series
df_grilla_resumen["OBSERVACION"] = empty_series
df_grilla_resumen["RECEPCIONADO"] = empty_series

#se agrupa por valores de jaula y patente dejando los nan de ultimo en cada patente
df_grilla_resumen = df_grilla_resumen.groupby("Patente")
df_grilla_resumen = df_grilla_resumen.apply(lambda x: x.sort_values("JAULA"))
df_grilla_resumen = df_grilla_resumen.reset_index(drop=True)
df_grilla_resumen["#"] = df_grilla_resumen.groupby("Patente").cumcount() +1


#MERGE DE LA INFORMACION DEL FLUJO CON LA GRILLA DE SODIMAC COMPLETA PARA SABER RECEPCIONADOS
df_grilla_resumen = pd.merge(df_grilla_resumen, df_flujo,on="LPN", how="left",indicator=True)
# Asignar los valores en la columna "RECEPCIONADO" según corresponda
df_grilla_resumen["RECEPCIONADO"] = np.where(df_grilla_resumen["_merge"] == "both", "SI", "")
# Eliminar la columna "_merge"
df_grilla_resumen = df_grilla_resumen.drop(columns=["_merge"])




#MERGE DE LA INFORMACION DE STOCK MUERTOS CON LA GRILLA DE SODIMAC COMPLETA PARA SABER RECEPCIONADOS
#mask = df_grilla_resumen["RECEPCIONADO"].isna()
#df_merge = pd.merge(df_grilla_resumen[mask], df_muertos, on="LPN", how="left")
#df_grilla_resumen.loc[mask, "RECEPCIONADO"] = df_merge["LPN"].apply(lambda x: "MUERTO" if pd.notna(x) else "")





# RECUENTO JAULAS UNICAS POR PATENTE
df_jaula = df_grilla_resumen.groupby('Patente')['JAULA'].unique().reset_index(name='JAULAS')

df_jaula = df_grilla_resumen[["Patente","JAULA"]]

df_jaula = df_jaula.dropna().copy()
df_jaula.loc[:, "JAULA"] = df_jaula["JAULA"].astype(int)
df_jaula = df_jaula.groupby("Patente")["JAULA"].unique().reset_index(name="JAULA")
df_jaula['JAULA'] = df_jaula['JAULA'].apply(lambda x: ', '.join(str(int(j)) for j in x))













# CUADRO RESUMEN PARA PATENTES, JAULAS QUE CONTIENEN, LPNS , DIRECCIONES UNICAS Y LOCALIDADES
df_final = df_direcciones.merge(df_localidades, on='Patente').merge(df_jaula, on='Patente').merge(df_count, on='Patente')
df_final = df_final.rename(columns={"Patente":"PATENTE","LPN":"LPN","DIRECCIONES UNICAS":"DIRECCIONES UNICAS","JAULA":"JAULA","LOCALIDADES":"LOCALIDADES"})

df_final = df_final.reindex(columns=["PATENTE", "DIRECCIONES UNICAS", "LPN", "JAULA", "LOCALIDADES"])
df_final.insert(0, "SHIPMENT", "")



#DF PARA GRILLA EN FORMATO TABLA EXTRAIDA DE LA GRILLA RESUMEN
df_definitiva = df_grilla_resumen[["#","JAU/STD/ID","RESER/SOC/SKU","COMUNA/PAT/EST","Patente","CLIENTE/TELEF","Direccion","LPN","Producto","CANTIDAD","RECEPCIONADO","V/M","BULTOS","OBSERVACION"]]





num_rows = len(df_definitiva) + 2  # sumar 1 para incluir el encabezado
ref = f"B2:O{num_rows}"

# crear un libro de trabajo de Excel y una hoja
book = Workbook()
sheet = book.active
sheet.title = "GRILLA"

# Agregar encabezados a la hoja de cálculo
for i, row in enumerate(dataframe_to_rows(df_definitiva, index=False, header=True), start=2):
    for j, value in enumerate(row, start=2):
        sheet.cell(row=i, column=j, value=value)

# Establecer el estilo de tabla
tab = Table(displayName="Table1", ref=ref)

style = TableStyleInfo(name="TableStyleMedium6", showFirstColumn=False,
                       showLastColumn=False, showRowStripes=True, showColumnStripes=True)

tab.tableStyleInfo = style

# Agregar tabla al archivo excel
sheet.add_table(tab)

# guardar el libro de trabajo de Excel
book.save(os.path.join(dir_actual, 'output_1.xlsx'))

#ELABORACION DEL MERGE ENTRE LA GRILLA DE FALABELLA Y LA DE SODIMAC
df_falabella = df_falabella.merge(df_wip[['JAULA',"LPN","BOD", 'BLOQUEADO',"CANTIDAD",'PESO','VOLUMEN_SODI',"SKU"]], on='LPN', how='left')
df_falabella["V/M"] = empty_series
df_falabella["BULTOS"] = empty_series
df_falabella["OBSERVACION"] = empty_series
df_falabella["RECEPCIONADO"] = empty_series

df_falabella = pd.merge(df_falabella, df_flujo,on="LPN", how="left",indicator=True)
# Asignar los valores en la columna "RECEPCIONADO" según corresponda
df_falabella["RECEPCIONADO"] = np.where(df_falabella["_merge"] == "both", "SI", "")
# Eliminar la columna "_merge"
df_falabella = df_falabella.drop(columns=["_merge"])




df_falabella = df_falabella[["JAULA","LPN","RESERVA","SKU","Producto","CANTIDAD","RECEPCIONADO","BOD","Patente","Direccion","Localidad","PESO","VOLUMEN_SODI","V/M","BULTOS","OBSERVACION"]]

#FILTRA EL DF FLUJO PARA RESTARLE LO QUE ESTÀ EN LAS GRILLAS DE FALABELLA Y SODIMAC, ESTOS PRODUCTOS PUEDEN ESTAR BLOQUEADOS, PUDIERON SIMPLEMENTE NO MIGRAR AL PANEL O
mask_primario = ~df_flujo['LPN'].isin(df_falabella['LPN']) & ~df_flujo['LPN'].isin(df_grilla_resumen['LPN'])
df_flujo_filtrado = df_flujo[mask_primario]
df_flujo_filtrado =df_flujo_filtrado.rename(columns={'OC': 'RESERVA'})
df_flujo_filtrado['RESERVA'] = df_flujo_filtrado['RESERVA'].astype(str)

#DE LO QUE QUEDA EN EL FLUJO Y NO EN LAS GRILLAS, SE FILTRA A VER QUE APARECE EN LA BASE DE DATOS
mask_db= df_flujo_filtrado['LPN'].isin(df_wip['LPN'])
df_flujo_en_db=df_flujo_filtrado[mask_db]
df_flujo_en_db = df_flujo_en_db.merge(df_wip[['JAULA',"LPN","BOD","PRODUCTO", 'BLOQUEADO',"CANTIDAD",'PESO','VOLUMEN_SODI',"SKU","DIRECCION","COMUNA"]], on='LPN', how='left')
df_flujo_en_db = df_flujo_en_db[["JAULA","ULTIMA_FECHA_ACT_UB","LPN","RESERVA","SKU","PRODUCTO","CANTIDAD","BOD","DIRECCION","COMUNA","PESO","VOLUMEN_SODI"]]

df_merged_sodimac = pd.merge(df_flujo_filtrado, df_sodimac, on='RESERVA', how='outer')
df_merged_sodimac = df_merged_sodimac[["JAULA","ULTIMA_FECHA_ACT_UB","LPN","RESERVA","SKU","PRODUCTO","CANTIDAD","DESP","ESTADO RESERVA","BLQ","DIRECCION","COMUNA","PESO","VOLUMEN_SODI"]]




#crear un archivo donde se muestra el flujo y lo que no se encuentra en las grillas
with pd.ExcelWriter(os.path.join(dir_actual, 'output_3.xlsx'), engine='openpyxl') as writer:
   df_flujo.to_excel(writer, sheet_name='lpn recepcionados', index=False)
   df_flujo_filtrado.to_excel(writer, sheet_name='total sin grilla', index=False)
   df_merged_sodimac.to_excel(writer, sheet_name='SIN GRILLA EN SODIMAC', index=False)
   df_flujo_en_db.to_excel(writer, sheet_name='aparecen en la db', index=False)



#crear otro archivo para la grilla de falabella
with pd.ExcelWriter(os.path.join(dir_actual, 'output_2.xlsx'), engine='openpyxl') as writer:
   
   #df_definitiva.to_excel(writer, sheet_name='grilla_modificada', index=False)
   df_falabella.to_excel(writer, sheet_name='falabella', index=False)




# Crear un nuevo archivo Excel y escribir los datos del dataframe en una nueva hoja de cálculo
with pd.ExcelWriter(os.path.join(dir_actual, 'output_1.xlsx'), engine='openpyxl', mode='a') as writer:
   
    #df_definitiva.to_excel(writer, sheet_name='grilla_modificada', index=False)
    df_wip.to_excel(writer, sheet_name='BASE_DE_DATOS', index=False)
    df_grilla_resumen.to_excel(writer, sheet_name='GRILLA_SIN_PROCESAR', index=False)
    df_flujo.to_excel(writer, sheet_name = "flujo", index=False)
    #df_merge.to_excel(writer,sheet_name="mask", index=False)

   
time.sleep(5)


       






# Crear un objeto Workbook
book = load_workbook(os.path.join(dir_actual, 'output_1.xlsx'))

# Crear una nueva hoja
muertos_sheet = book.create_sheet('BUSCADOR_MUERTOS_GRILLA')





# ESCRIBE EN BUSCADOR_MUERTOS_GRILLA LOS BUSCARV CON LA INFOMRACION NECESARIA EXTRAIDA DE GRILLA_SIN_PROCESAR
muertos_sheet['A1'] = 'LPN'
muertos_sheet['B1'] = 'JAU/STD/ID'
muertos_sheet['C1'] = 'RESER/SOC/SKU'
muertos_sheet['D1'] = 'COMUNA/PAT/EST'
muertos_sheet["E1"] = "PATENTE"  
muertos_sheet["F1"] = "PRODUCTO" 
muertos_sheet["G1"] = "CANTIDAD" 


for i in range(2, 50):
    muertos_sheet['B' + str(i)] = '=VLOOKUP(A' + str(i) + ',GRILLA_SIN_PROCESAR!B:V,21,FALSE)'
    muertos_sheet['B' + str(i)].data_type = 'f'

    muertos_sheet['C' + str(i)] = '=VLOOKUP(A' + str(i) + ',GRILLA_SIN_PROCESAR!B:W,22,FALSE)'
    muertos_sheet['C' + str(i)].data_type = 'f'
  
    muertos_sheet['D' + str(i)] = '=VLOOKUP(A' + str(i) + ',GRILLA_SIN_PROCESAR!B:X,23,FALSE)'
    muertos_sheet['D' + str(i)].data_type = 'f'

    muertos_sheet['E' + str(i)] = '=VLOOKUP(A' + str(i) + ',GRILLA_SIN_PROCESAR!B:H,7,FALSE)'
    muertos_sheet['E' + str(i)].data_type = 'f'
    
    muertos_sheet['F' + str(i)] = '=VLOOKUP(A' + str(i) + ',GRILLA_SIN_PROCESAR!B:E,4,FALSE)'
    muertos_sheet['F' + str(i)].data_type = 'f'

    muertos_sheet['G' + str(i)] = '=VLOOKUP(A' + str(i) + ',GRILLA_SIN_PROCESAR!B:M,12,FALSE)'
    muertos_sheet['G' + str(i)].data_type = 'f'




# Crear una nueva hoja
muertos_sheet = book.create_sheet('ANOMALIAS')

muertos_sheet['A1'] = 'LPN'
muertos_sheet['B1'] = 'HUB'
muertos_sheet['C1'] = 'RESERVA'
muertos_sheet['D1'] = 'ORIGEN'
muertos_sheet["E1"] = "SKU"  
muertos_sheet["F1"] = "DESCRIPCION" 
muertos_sheet["G1"] = "CANTIDAD"
muertos_sheet["H1"] = "ANOMALIA DETECTADA"

for i in range(2, 50):
  

    muertos_sheet['B' + str(i)] = '=VLOOKUP(A' + str(i) + ',BASE_DE_DATOS!A:K,11,FALSE)'
    muertos_sheet['B' + str(i)].data_type = 'f'

    muertos_sheet['C' + str(i)] = '=VLOOKUP(A' + str(i) + ',GRILLA_SIN_PROCESAR!B:D,3,FALSE)'
    muertos_sheet['C' + str(i)].data_type = 'f'

    muertos_sheet['D' + str(i)] = '=VLOOKUP(A' + str(i) + ',GRILLA_SIN_PROCESAR!B:O,14,FALSE)'
    muertos_sheet['D' + str(i)].data_type = 'f'

    muertos_sheet['E' + str(i)] = '=VLOOKUP(A' + str(i) + ',GRILLA_SIN_PROCESAR!B:L,11,FALSE)'
    muertos_sheet['E' + str(i)].data_type = 'f'

    muertos_sheet['F' + str(i)] = '=VLOOKUP(A' + str(i) + ',GRILLA_SIN_PROCESAR!B:E,4,FALSE)'
    muertos_sheet['F' + str(i)].data_type = 'f'

    muertos_sheet['G' + str(i)] = '=VLOOKUP(A' + str(i) + ',GRILLA_SIN_PROCESAR!B:M,12,FALSE)'
    muertos_sheet['G' + str(i)].data_type = 'f'

    muertos_sheet['H' + str(i)] = '=VLOOKUP(A' + str(i) + ',GRILLA!I:N,6,FALSE)'
    muertos_sheet['H' + str(i)].data_type = 'f'



muertos_sheet = book.create_sheet('BUSCADOR_EN_DB_Y_AGREGADOS')

muertos_sheet["B2"]="LPN"
muertos_sheet["C2"]="RESERVA"
muertos_sheet["D2"]="SOC"
muertos_sheet["E2"]="SKU"
muertos_sheet["F2"]="PRODUCTO"
muertos_sheet["G2"]="CANTIDAD"
muertos_sheet["H2"]="DIRECCION"
muertos_sheet["I2"]="LOCALIDAD"
muertos_sheet["J2"]="ID"
muertos_sheet["K2"]="PATENTE"
muertos_sheet["L2"]="BULTOS"
muertos_sheet["M2"]="OBSERVACION"

for i in range(3, 50):
    muertos_sheet['C' + str(i)] = '=VLOOKUP(B' + str(i) + ',BASE_DE_DATOS!A:C,3,FALSE)'
    muertos_sheet['C' + str(i)].data_type = 'f'

    muertos_sheet['D' + str(i)] = '=VLOOKUP(B' + str(i) + ',BASE_DE_DATOS!A:H,8,FALSE)'
    muertos_sheet['D' + str(i)].data_type = 'f'

    muertos_sheet['E' + str(i)] = '=VLOOKUP(B' + str(i) + ',BASE_DE_DATOS!A:D,4,FALSE)'
    muertos_sheet['E' + str(i)].data_type = 'f'

    muertos_sheet['F' + str(i)] = '=VLOOKUP(B' + str(i) + ',BASE_DE_DATOS!A:E,5,FALSE)'
    muertos_sheet['F' + str(i)].data_type = 'f'

    muertos_sheet['G' + str(i)] = '=VLOOKUP(B' + str(i) + ',BASE_DE_DATOS!A:F,6,FALSE)'
    muertos_sheet['G' + str(i)].data_type = 'f'

    muertos_sheet['H' + str(i)] = '=VLOOKUP(B' + str(i) + ',BASE_DE_DATOS!A:R,18,FALSE)'
    muertos_sheet['H' + str(i)].data_type = 'f'

    muertos_sheet['I' + str(i)] = '=VLOOKUP(B' + str(i) + ',BASE_DE_DATOS!A:Q,17,FALSE)'
    muertos_sheet['I' + str(i)].data_type = 'f'


# Crear una nueva hoja, devoluciones
devoluciones_sheet = book.create_sheet('DEVOLUCIONES')


devoluciones_sheet['A1'] = 'MES'
devoluciones_sheet['B1'] = 'FECHA'
devoluciones_sheet['C1'] = 'ESTADO'
devoluciones_sheet['D1'] = 'PALLET'
devoluciones_sheet["E1"] = "RESERVA H"  
devoluciones_sheet["F1"] = "RESERVA M" 
devoluciones_sheet["G1"] = "SKU" 
devoluciones_sheet["H1"] = "Descripcion" 
devoluciones_sheet["I1"] = "CAN" 
devoluciones_sheet["J1"] = "LPN/POS/RD" 
devoluciones_sheet["K1"] = "LPN" 
devoluciones_sheet["L1"] = "DAÑO PRODUCTO" 
devoluciones_sheet["M1"] = "DAÑO EMBALAJE" 
devoluciones_sheet["N1"] = "EMPRESA" 
devoluciones_sheet["O1"] = "TABLA" 
devoluciones_sheet["P1"] = "BULTO" 
devoluciones_sheet["Q1"] = "TRINQUETE" 

for i in range(2, 60):
    devoluciones_sheet['F' + str(i)] = '=VLOOKUP(K' + str(i) + ',BASE_DE_DATOS!A:B,2,FALSE)'
    devoluciones_sheet['F' + str(i)].data_type = 'f'

    devoluciones_sheet['G' + str(i)] = '=VLOOKUP(K' + str(i) + ',BASE_DE_DATOS!A:I,9,FALSE)'
    devoluciones_sheet['G' + str(i)].data_type = 'f'

    devoluciones_sheet['H' + str(i)] = '=VLOOKUP(K' + str(i) + ',BASE_DE_DATOS!A:K,11,FALSE)'
    devoluciones_sheet['H' + str(i)].data_type = 'f'
    
    devoluciones_sheet['I' + str(i)] = '=VLOOKUP(K' + str(i) + ',BASE_DE_DATOS!A:J,10,FALSE)'
    devoluciones_sheet['I' + str(i)].data_type = 'f'



crear_hoja = book.create_sheet('TIM')

#for r in dataframe_to_rows(df_final, index=False, header=True):
 #   crear_hoja.append(r)


for r in dataframe_to_rows(df_final, index=False, header=True):
    crear_hoja.append(r)


crear_hoja = book.create_sheet('AGREGADOS')

crear_hoja['B2'] = 'ID'
crear_hoja['C2'] = 'SOC'
crear_hoja['D2'] = 'RESERVA'
crear_hoja['E2'] = 'COMUNA'
crear_hoja["F2"] = "PATENTE"  
crear_hoja["G2"] = "CLIENTE" 
crear_hoja["H2"] = "DIRECCION" 
crear_hoja["I2"] = "LPN" 
crear_hoja["J2"] = "PRODUCTO" 
crear_hoja["K2"] = "CANTIDAD" 
crear_hoja["L2"] = "V/M" 
crear_hoja["M2"] = "BULTO" 
crear_hoja["N2"] = "OBSERVACION" 

for i in range(2, 60):
    crear_hoja['D' + str(i)] = '=VLOOKUP(I' + str(i) + ',BASE_DE_DATOS!A:C,3,FALSE)'
    crear_hoja['D' + str(i)].data_type = 'f'

    crear_hoja['E' + str(i)] = '=VLOOKUP(I' + str(i) + ',BASE_DE_DATOS!A:Q,17,FALSE)'
    crear_hoja['E' + str(i)].data_type = 'f'

    crear_hoja['G' + str(i)] = '=VLOOKUP(I' + str(i) + ',BASE_DE_DATOS!A:S,19,FALSE)'
    crear_hoja['G' + str(i)].data_type = 'f'
    
    crear_hoja['H' + str(i)] = '=VLOOKUP(I' + str(i) + ',BASE_DE_DATOS!A:R,18,FALSE)'
    crear_hoja['H' + str(i)].data_type = 'f'

    crear_hoja['J' + str(i)] = '=VLOOKUP(I' + str(i) + ',BASE_DE_DATOS!A:E,5,FALSE)'
    crear_hoja['J' + str(i)].data_type = 'f'

    crear_hoja['K' + str(i)] = '=VLOOKUP(I' + str(i) + ',BASE_DE_DATOS!A:F,6,FALSE)'
    crear_hoja['K' + str(i)].data_type = 'f'

    crear_hoja['C' + str(i)] = '=VLOOKUP(I' + str(i) + ',BASE_DE_DATOS!A:H,8,FALSE)'
    crear_hoja['C' + str(i)].data_type = 'f'


book.save(os.path.join(dir_actual, 'output_1.xlsx'))

































